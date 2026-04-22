from flask import Flask, redirect, render_template, request, url_for, jsonify, send_file, flash, send_from_directory
from flask_sqlalchemy import SQLAlchemy
from flask_login import LoginManager, UserMixin, login_user, logout_user, login_required, current_user
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime
from zoneinfo import ZoneInfo
import uuid
import os
import io
import csv
import base64
import qrcode
import re

basedir = os.path.abspath(os.path.dirname(__file__))

app = Flask(__name__)
app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///' + \
    os.path.join(basedir, 'tiket.db')
app.config['SECRET_KEY'] = 'mms-kajian-secret-key-2026'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
db = SQLAlchemy(app)

WIB = ZoneInfo('Asia/Jakarta')


def wib_now():
    return datetime.now(WIB)


def format_wib(dt):
    if dt is None:
        return '-'
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=WIB)
    else:
        dt = dt.astimezone(WIB)
    return dt.strftime('%d/%m/%Y %H:%M WIB')


@app.template_filter('wib')
def wib_filter(dt):
    return format_wib(dt)


def sanitize_input(text, max_length=100):
    if not text:
        return ''
    cleaned = re.sub(r'<[^>]+>', '', text)
    cleaned = re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', cleaned)
    return cleaned.strip()[:max_length]


login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Silakan login terlebih dahulu.'
login_manager.login_message_category = 'warning'


class User(UserMixin, db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)


class Tiket(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    kode = db.Column(db.String(10), unique=True, nullable=False)
    nama = db.Column(db.String(100), default="Peserta Kajian MMS")
    angkatan = db.Column(db.String(50))
    is_used = db.Column(db.Boolean, default=False)
    waktu_daftar = db.Column(db.DateTime, default=wib_now)
    waktu_scan = db.Column(db.DateTime, nullable=True)


class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True, nullable=False)
    value = db.Column(db.String(200), nullable=False)


@login_manager.user_loader
def load_user(user_id):
    return User.query.get(int(user_id))


def get_kuota():
    setting = Setting.query.filter_by(key='kuota').first()
    return int(setting.value) if setting else 70

# UMUM


def get_kuota_umum():
    setting = Setting.query.filter_by(key='kuota_umum').first()
    return int(setting.value) if setting else 50


def generate_qr_base64(data):
    img = qrcode.make(data)
    buffer = io.BytesIO()
    img.save(buffer, format='PNG')
    buffer.seek(0)
    return base64.b64encode(buffer.getvalue()).decode('utf-8')


@app.route('/login', methods=['GET', 'POST'])
def login():
    if current_user.is_authenticated:
        return redirect(url_for('admin'))
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        user = User.query.filter_by(username=username).first()
        if user and user.check_password(password):
            login_user(user)
            next_page = request.args.get('next')
            return redirect(next_page or url_for('admin'))
        flash('Username atau password salah!', 'danger')
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    logout_user()
    flash('Berhasil logout.', 'info')
    return redirect(url_for('login'))


@app.route('/')
def home():
    if not current_user.is_authenticated:
        return redirect(url_for('halaman_daftar'))
    return render_template('index.html', status=None)


@app.route('/scan/<kode_tiket>')
@login_required
def scan(kode_tiket):
    tiket = Tiket.query.filter_by(kode=kode_tiket).first()
    if not tiket:
        return render_template('index.html', status='tidak ditemukan', kode=kode_tiket)
    if tiket.is_used:
        return render_template('index.html', status='telah terpakai', kode=kode_tiket,
                               nama_peserta=tiket.nama, angkatan_peserta=tiket.angkatan)
    else:
        tiket.is_used = True
        tiket.waktu_scan = wib_now()
        db.session.commit()
        return render_template('index.html', status='berhasil', kode=kode_tiket,
                               nama_peserta=tiket.nama, angkatan_peserta=tiket.angkatan)


@app.route('/cek_manual', methods=['POST'])
@login_required
def cek_manual():
    kode = request.form.get('kode_input', '').strip()
    if kode:
        return redirect(url_for('scan', kode_tiket=kode))
    return redirect(url_for('home'))


@app.route('/pendaftaran')
def halaman_depan():
    return render_template('pendaftaran.html')


@app.route('/daftar')
def halaman_daftar():
    kuota = get_kuota()
    total_terdaftar = Tiket.query.count()
    sisa_kuota = max(0, kuota - total_terdaftar)
    if total_terdaftar >= kuota:
        return render_template('habis.html', kuota=kuota)
    return render_template('daftar.html', sisa_kuota=sisa_kuota, total_terdaftar=total_terdaftar, kuota=kuota)

# INI UMUM JUGA


@app.route('/umum_daftar')
def halaman_umum_daftar():
    return render_template('umum_daftar.html')


@app.route('/proses_daftar', methods=['POST'])
def proses_daftar():
    kuota = get_kuota()
    total_terdaftar = Tiket.query.count()
    if total_terdaftar >= kuota:
        return render_template('habis.html', kuota=kuota)

    nama = sanitize_input(request.form.get('nama_input', ''))
    pilihan_kelas = sanitize_input(request.form.get('kelas_input', ''))

    if not nama or not pilihan_kelas:
        flash('Nama dan angkatan wajib diisi!', 'danger')
        return redirect(url_for('halaman_daftar'))

    kode_otomatis = "MMS-" + str(uuid.uuid4()).upper()[:4]

    peserta_baru = Tiket(nama=nama, angkatan=pilihan_kelas, kode=kode_otomatis)
    db.session.add(peserta_baru)
    db.session.commit()

    qr_base64 = generate_qr_base64(kode_otomatis)

    return render_template('sukses.html', nama=nama, angkatan=pilihan_kelas,
                           kode=kode_otomatis, qr_base64=qr_base64)


# INI PROSES DAFTAR UMUM
@app.route('/proses_daftar_umum', methods=['POST'])
def proses_daftar_umum():
    kuota = get_kuota_umum()
    total_terdaftar = Tiket.query.count()
    if total_terdaftar >= kuota:
        return render_template('habis.html', kuota=kuota)

    nama = sanitize_input(request.form.get('nama_input_umum', ''))
    pilihan_umum = sanitize_input(request.form.get('kelas_input_umum', ''))

    if not nama or not pilihan_umum:
        flash('Nama dan angkatan wajib diisi!', 'danger')
        return redirect(url_for('halaman_umum_daftar'))

    kode_otomatis = "MMS-" + str(uuid.uuid4()).upper()[:4]

    peserta_baru = Tiket(
        nama_umum=nama, angkatan_umum=pilihan_umum, kode=kode_otomatis)
    db.session.add(peserta_baru)
    db.session.commit()

    qr_base64 = generate_qr_base64(kode_otomatis)

    return render_template('sukses.html', nama_umum=nama, angkatan_umum=pilihan_umum,
                           kode=kode_otomatis, qr_base64=qr_base64)


@app.route('/cek', methods=['GET', 'POST'])
def cek_tiket():
    result = None
    kode_input = ''
    if request.method == 'POST':
        kode_input = request.form.get('kode_input', '').strip()
        tiket = Tiket.query.filter_by(kode=kode_input).first()
        if tiket:
            result = {
                'found': True,
                'nama': tiket.nama if hasattr(tiket, 'nama') else tiket.nama_umum,
                'angkatan': tiket.angkatan if hasattr(tiket, 'angkatan') else tiket.angkatan_umum,
                'kode': tiket.kode,
                'is_used': tiket.is_used,
                'waktu_daftar': format_wib(tiket.waktu_daftar),
                'waktu_scan': format_wib(tiket.waktu_scan)
            }
        else:
            result = {'found': False}
    return render_template('cek_tiket.html', result=result, kode_input=kode_input)


@app.route('/form')
def form_data_diri():
    return render_template('form.html')


@app.route('/form_umum')
def form_umum():
    return render_template('form_umum.html')


@app.route('/whatsapp')
def redirect_whatsapp():
    return render_template('whatsapp.html')


@app.route('/admin')
@login_required
def admin():
    kuota = get_kuota()
    semua_tiket = Tiket.query.order_by(Tiket.waktu_daftar.desc()).all()
    total = len(semua_tiket)
    terpakai = sum(1 for t in semua_tiket if t.is_used)
    sisa = total - terpakai
    return render_template('admin.html', tiket=semua_tiket, total=total,
                           terpakai=terpakai, sisa=sisa, kuota=kuota)


@app.route('/admin/kuota', methods=['POST'])
@login_required
def set_kuota():
    new_kuota = request.form.get('kuota', '70').strip()
    try:
        new_kuota = int(new_kuota)
        if new_kuota < 1:
            flash('Kuota minimal 1!', 'danger')
            return redirect(url_for('admin'))
    except ValueError:
        flash('Kuota harus berupa angka!', 'danger')
        return redirect(url_for('admin'))

    setting = Setting.query.filter_by(key='kuota').first()
    if setting:
        setting.value = str(new_kuota)
    else:
        setting = Setting(key='kuota', value=str(new_kuota))
        db.session.add(setting)
    db.session.commit()
    flash(f'Kuota pendaftaran berhasil diubah menjadi {new_kuota}.', 'success')
    return redirect(url_for('admin'))


@app.route('/admin/walkin', methods=['POST'])
@login_required
def walkin_register():
    nama = sanitize_input(request.form.get('nama', ''))
    angkatan = sanitize_input(request.form.get('angkatan', ''))

    if not nama or not angkatan:
        flash('Nama dan angkatan wajib diisi untuk walk-in!', 'danger')
        return redirect(url_for('admin'))

    kode = "WLK-" + str(uuid.uuid4()).upper()[:4]
    now = wib_now()
    peserta = Tiket(nama=nama, angkatan=angkatan, kode=kode,
                    is_used=True, waktu_daftar=now, waktu_scan=now)
    db.session.add(peserta)
    db.session.commit()
    flash(
        f'Walk-in berhasil! {nama} ({angkatan}) terdaftar dengan kode {kode}.', 'success')
    return redirect(url_for('admin'))


@app.route('/reset_semua', methods=['POST'])
@login_required
def reset_semua():
    Tiket.query.update({Tiket.is_used: False, Tiket.waktu_scan: None})
    db.session.commit()
    flash('Semua tiket berhasil direset!', 'success')
    return redirect(url_for('admin'))


@app.route('/reset_tiket/<int:tiket_id>', methods=['POST'])
@login_required
def reset_tiket(tiket_id):
    tiket = Tiket.query.get_or_404(tiket_id)
    tiket.is_used = False
    tiket.waktu_scan = None
    db.session.commit()
    flash(f'Tiket {tiket.kode} berhasil direset.', 'success')
    return redirect(url_for('admin'))


@app.route('/hapus_tiket/<int:tiket_id>', methods=['POST'])
@login_required
def hapus_tiket(tiket_id):
    tiket = Tiket.query.get_or_404(tiket_id)
    kode = tiket.kode
    db.session.delete(tiket)
    db.session.commit()
    flash(f'Tiket {kode} berhasil dihapus.', 'success')
    return redirect(url_for('admin'))


@app.route('/export/csv')
@login_required
def export_csv():
    tikets = Tiket.query.order_by(Tiket.waktu_daftar.desc()).all()
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(['No', 'Nama', 'Angkatan', 'Kode Tiket',
                    'Status', 'Waktu Daftar', 'Waktu Scan'])
    for i, t in enumerate(tikets, 1):
        writer.writerow([
            i, t.nama, t.angkatan, t.kode,
            'Hadir' if t.is_used else 'Belum Hadir',
            t.waktu_daftar.strftime(
                '%d/%m/%Y %H:%M') if t.waktu_daftar else '-',
            t.waktu_scan.strftime('%d/%m/%Y %H:%M') if t.waktu_scan else '-'
        ])
    output.seek(0)
    return send_file(
        io.BytesIO(output.getvalue().encode('utf-8-sig')),
        mimetype='text/csv',
        as_attachment=True,
        download_name=f'rekap_tiket_mms_{datetime.now().strftime("%Y%m%d")}.csv'
    )


@app.route('/export/excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()
    ws = wb.active
    ws.title = "Rekap Tiket MMS"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6B4C7A",
                              end_color="6B4C7A", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['No', 'Nama', 'Angkatan', 'Kode Tiket',
               'Status', 'Waktu Daftar', 'Waktu Scan']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    tikets = Tiket.query.order_by(Tiket.waktu_daftar.desc()).all()
    for i, t in enumerate(tikets, 1):
        row = i + 1
        ws.cell(row=row, column=1, value=i).border = thin_border
        ws.cell(row=row, column=2, value=t.nama).border = thin_border
        ws.cell(row=row, column=3, value=t.angkatan).border = thin_border
        ws.cell(row=row, column=4, value=t.kode).border = thin_border
        ws.cell(row=row, column=5,
                value='Hadir' if t.is_used else 'Belum Hadir').border = thin_border
        ws.cell(row=row, column=6, value=t.waktu_daftar.strftime(
            '%d/%m/%Y %H:%M') if t.waktu_daftar else '-').border = thin_border
        ws.cell(row=row, column=7, value=t.waktu_scan.strftime(
            '%d/%m/%Y %H:%M') if t.waktu_scan else '-').border = thin_border

        status_cell = ws.cell(row=row, column=5)
        if t.is_used:
            status_cell.fill = PatternFill(
                start_color="D4EDDA", end_color="D4EDDA", fill_type="solid")
        else:
            status_cell.fill = PatternFill(
                start_color="FFF3CD", end_color="FFF3CD", fill_type="solid")

    for col in range(1, 8):
        max_len = max(len(str(ws.cell(row=r, column=col).value or ''))
                      for r in range(1, len(tikets) + 2))
        ws.column_dimensions[chr(64 + col)].width = max(max_len + 4, 12)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'rekap_tiket_mms_{datetime.now().strftime("%Y%m%d")}.xlsx'
    )


@app.route('/export/pdf')
@login_required
def export_pdf():
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.units import mm

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=landscape(
        A4), topMargin=20*mm, bottomMargin=20*mm)
    elements = []
    styles = getSampleStyleSheet()

    title = Paragraph("Rekap Tiket Kajian Offline MMS 2026", styles['Title'])
    elements.append(title)
    elements.append(Spacer(1, 10*mm))

    tikets = Tiket.query.order_by(Tiket.waktu_daftar.desc()).all()
    data = [['No', 'Nama', 'Angkatan', 'Kode Tiket',
             'Status', 'Waktu Daftar', 'Waktu Scan']]
    for i, t in enumerate(tikets, 1):
        data.append([
            str(i), t.nama, t.angkatan, t.kode,
            'Hadir' if t.is_used else 'Belum Hadir',
            t.waktu_daftar.strftime(
                '%d/%m/%Y %H:%M') if t.waktu_daftar else '-',
            t.waktu_scan.strftime('%d/%m/%Y %H:%M') if t.waktu_scan else '-'
        ])

    col_widths = [30, 140, 70, 80, 70, 100, 100]
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#6B4C7A')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor('#F9F0F7')]),
    ]))
    elements.append(table)

    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'rekap_tiket_mms_{datetime.now().strftime("%Y%m%d")}.pdf'
    )


@app.route('/api/stats')
@login_required
def api_stats():
    tikets = Tiket.query.all()
    kuota = get_kuota()

    angkatan_data = {}
    for t in tikets:
        ang = t.angkatan or 'Lainnya'
        if ang not in angkatan_data:
            angkatan_data[ang] = {'total': 0, 'hadir': 0}
        angkatan_data[ang]['total'] += 1
        if t.is_used:
            angkatan_data[ang]['hadir'] += 1

    timeline = {}
    for t in tikets:
        if t.waktu_scan:
            hour_key = t.waktu_scan.strftime('%H:%M')
            timeline[hour_key] = timeline.get(hour_key, 0) + 1

    sorted_timeline = dict(sorted(timeline.items()))

    total = len(tikets)
    hadir = sum(1 for t in tikets if t.is_used)

    tikets_ordered = Tiket.query.order_by(Tiket.waktu_daftar.desc()).all()

    return jsonify({
        'total': total,
        'hadir': hadir,
        'belum_hadir': total - hadir,
        'kuota': kuota,
        'sisa_kuota': max(0, kuota - total),
        'angkatan': angkatan_data,
        'timeline': sorted_timeline,
        'tikets': [{
            'id': t.id,
            'nama': t.nama,
            'angkatan': t.angkatan,
            'kode': t.kode,
            'is_used': t.is_used,
            'waktu_daftar': format_wib(t.waktu_daftar),
            'waktu_scan': format_wib(t.waktu_scan)
        } for t in tikets_ordered]
    })


@app.route('/images/<path:filename>')
def serve_image(filename):
    return send_from_directory(os.path.join(basedir, 'images'), filename)


with app.app_context():
    db.create_all()

    if not User.query.first():
        admin_user = User(username='admin')
        admin_user.set_password('admin123')
        db.session.add(admin_user)
        db.session.commit()
        print("[OK] Default admin created (admin/admin123)")

    if not Setting.query.filter_by(key='kuota').first():
        db.session.add(Setting(key='kuota', value='70'))
        db.session.commit()


if __name__ == '__main__':
    app.run(host='0.0.0.0', port=5000, debug=True)
